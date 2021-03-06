from django.shortcuts import render
from django.http import HttpResponseRedirect
from .forms import FormularioNombre
from django.urls import reverse
from .models import Datosbanesco
from django.core.files.storage import FileSystemStorage
from django.conf import settings
import os
import openpyxl
from django.http import HttpResponse


def index(request):
    if request.method == "POST":
        form = FormularioNombre(request.POST, request.FILES)
        if form.is_valid:
            form.save()
            nombre = request.FILES['archivo'].name
            return HttpResponseRedirect(reverse('grabar', args=(nombre,)))
            
            
    else:
        form = FormularioNombre()
    return render(request, 'banesco/index.html', {'form': form})


#def index(request):
#    if request.method == "POST":
#        form = FormularioNombre(request.POST)
#        if form.is_valid:
#            nombre = request.POST['nombre_archivo']
#            uploaded_file = request.FILES['archivo']
#            fs = FileSystemStorage()
#            fs.save(uploaded_file.name,uploaded_file)
#            return HttpResponseRedirect(reverse('grabar', args=(nombre,)))
#    else:
#        form = FormularioNombre()
#    return render(request, 'banesco/index.html', {'form': form})


def grabar(request,nombre):
    ruta = os.path.join(os.path.join(settings.MEDIA_ROOT,'banescofiles'),nombre)
    print(ruta)
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="xxx.xlsx"'

    def concoma(valor):
        if valor.strip() == "":
            return ""
        else:
            largo = len(valor)
            entero = valor[0:largo-2]
            decimal = valor[largo-2:largo]
            return float(entero + "." + decimal)

    wb = openpyxl.Workbook()
    hoja = wb.get_sheet_by_name('Sheet')
    titulo = ['INICIO','REFPAG','MEDPAG','DIAPAG1', 'MESPA1','SIGPA1','AÑOPA1','DIGITO', 'MONFACT', 'NIAFAC', 'SIGFAC', 'AÑO', 'MES', 'DIG', 'MONTO']

    j = 1
    for k in titulo:
        hoja.cell(row=1,column=j).value = k
        j+= 1

    archivo = open(ruta,'r')
    linea = archivo.readline()
    contador = 1

    while linea != "":

        if linea[0:1] != "1":
            INICIO = linea[0:5]
            REFPAG = linea[5:16]
            MEDPAG = linea[16:19]
            DIAPAG1 = linea[19:21]
            MESPA1 = linea[21:23]
            SIGPA1 = linea[23:25]
            ANOPA1 = linea[25:27]
            DIGITO = linea[27:28]
            MONFACT = linea[28:47]
            NIAFAC = linea[47:55]
            SIGFAC = linea[55:57]
            ANO = linea[57:59]
            MES = linea[59:61]
            DIG = linea[61:62]
            
            hoja.cell(row=contador,column=1).value = INICIO
            hoja.cell(row=contador,column=2).value = REFPAG
            hoja.cell(row=contador,column=3).value = MEDPAG
            hoja.cell(row=contador,column=4).value = DIAPAG1
            hoja.cell(row=contador,column=5).value = MESPA1
            hoja.cell(row=contador,column=6).value = SIGPA1
            hoja.cell(row=contador,column=7).value = ANOPA1
            hoja.cell(row=contador,column=8).value = DIGITO
            hoja.cell(row=contador,column=9).value = MONFACT
            hoja.cell(row=contador,column=10).value = NIAFAC
            hoja.cell(row=contador,column=11).value = SIGFAC
            hoja.cell(row=contador,column=12).value = ANO
            hoja.cell(row=contador,column=13).value = MES
            hoja.cell(row=contador,column=14).value = DIG
            hoja.cell(row=contador,column=15).value = concoma(MONFACT)    
            
        linea = archivo.readline()
        contador+=1
        
    archivo.close()
    wb.save(response)
    return response

def vistaprueba(request):
    #familia = {'papa': 'leo', 'mama': 'yuli', 'hijo':'andre'}
    return render(request, 'banesco/pagina.html')



